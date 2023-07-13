# app file
import pandas as pd
import sys, requests, json, csv
from prettytable import PrettyTable
from openpyxl import load_workbook
#this new
def main():
    # TODO: Check for command-line usage
    # while True:
    #    try:
    #        if len(sys.argv) == 3:
    #            break
    #    except ValueError:
    #        print("dna.py, database .CSV path, DNA sequence to identify .txt path!")
    #        sys.exit(1)
    acs=pd.read_acsv("acs.csv", index_col=0)

    MWdict = {}
    for entry in acs["AC"]:
        MWdict[entry] = uniprot(entry)[0]
    molweights, MWlist = mw_extract()
    weights = pd.Series(molweights, index=acs["AC"], name="MW")

    table = pd.DataFrame(weights, index=acs["AC"])

    table.insert(1, column="MW uniprot", value=MWdict)

    fits = pd.Series(mw_copmarison(molweights, acs["AC"]), index=acs["AC"], name="MW fit")
    table.insert(2, column="Fit", value=fits)

    output = table.to_string(formatters={'Fit' : '{:,.1%}'.format})
    print(output)

    #table.to_csv("output.csv")

def ac_list():
    ac = []
    with open("acs.csv", "r") as lane_acs:
        reader = csv.reader(lane_acs)
        next(reader)
        for row in reader:
            ac.append(row[1])

    return ac


def mw_extract():
    molweights = []
    MWlist = []
    wb = load_workbook(filename="gel1.xlsx")

    # Number of the row, where the first MW entry is located, typically, in row 4
    rownum = 4

    # Loop through each well (every 9th column) to extract only the MW and omit the rest of parameters in a lane
    #MWlist.append({"Lane 1": None})
    for i in range(12, 229, 9):
        # Loop through top8 bands (rows 4 to 12) and find the thickest band (largest volume at +2 column)
        Thickest = 0
        if i == 120:
            #MWlist.append({"Lane 14": None})
            continue  # skips lane 14 (protein ladder)
        for j in range(4, 12):
            Volume = wb["Sheet1"].cell(row=j, column=(i + 2)).value
            if (Volume != None) and (Volume > Thickest):
                Thickest = Volume
                # Remember the row number of the thickest band
                rownum = j

        # copy MW of the thickest band to a list
        band = wb["Sheet1"].cell(row=rownum, column=i).value
        lane = wb["Sheet1"].cell(row=2, column = i - 2).value
        if band is None:
            MWdict = {lane: band}
            molweights.append(band)
        else:
            MWdict = {lane: float(band)}
            molweights.append(float(band))
        MWlist.append(MWdict)
    return molweights, MWlist


def uniprot(accession):
    #accession = "P01583"

    WEBSITE_API = "https://rest.uniprot.org/"

    def get_url(url, **kwargs):
        response = requests.get(url, **kwargs)

        if not response.ok:
            print(response.text)
            response.raise_for_status()
            sys.exit()

        return response

    ecodict = {
        "ECO:0000269": "EXP: Inferred from Experiment",
        "ECO:0000314": "IDA: Inferred from Direct Assay",
        "ECO:0000353": "IPI: Inferred from Physical Interaction",
        "ECO:0000315": "IMP: Inferred from Mutant Phenotype",
        "ECO:0000316": "IGI: Inferred from Genetic Interaction",
        "ECO:0000270": "IEP: Inferred from Expression Pattern",
        "ECO:0000250": "ISS: Inferred from Sequence or Structural Similarity",
        "ECO:0000255": "ISM: Inferred from Sequence Model",
    }

    r = get_url(
        f"{WEBSITE_API}/uniprotkb/{accession}?fields=ft_carbohyd&fields=ft_mod_res&fields=ft_lipid&fields=mass"
    )
    data = r.json()
    while True:
        try:
            if data["sequence"]["molWeight"] > 1:
                break
        except KeyError:
            print(f"{accession} no uniprot entry found")
            return None, None

    MWest = data["sequence"]["molWeight"]/1000

    PTMlist = []
    for feature in data['features']:
        PTMdict = {
            "type": feature["type"],
            "description": feature["description"],
            "location": feature["location"]["start"]["value"],
            "evidence": ecodict[feature["evidences"][0]["evidenceCode"]],
        }
        PTMlist.append(PTMdict)
    return (MWest, PTMlist)


def mw_copmarison(estimated, ac):
    mw = []
    fits = []
    with open("databaseMW.csv", "r") as db:
        reader = csv.reader(db)
        for row in reader:
            for i in ac:
                if row[0] == i:
                    mw.append(float(row[1]))
    j = 0
    for i in mw:
        if estimated[j] == None:
            fit = None
        elif estimated[j] >= i:
            fit = (estimated[j] - (estimated[j] - i)) / estimated[j]
        elif estimated[j] < i:
            fit = (i - (i - estimated[j])) / i
        else:
            sys.exit("Failed to calculate MW fit")
        fits.append(fit)
        j += 1

    return fits



if __name__ == "__main__":
    main()
